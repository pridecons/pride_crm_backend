from __future__ import annotations
import csv
import io
import json
import os
from typing import Optional, List, Dict, Any, Tuple, Set
from fastapi import APIRouter, Depends, HTTPException, status, File, UploadFile, Form
from sqlalchemy.orm import Session
from sqlalchemy import insert, select, func
from pydantic import BaseModel
from routes.auth.auth_dependency import get_current_user
from db.connection import get_db
from db.models import Lead, LeadSource, UserDetails
from utils.validation_utils import FormatValidator

# NEW: Excel support
try:
    import openpyxl  # type: ignore
except Exception:  # pragma: no cover
    openpyxl = None  # we'll error clearly if user uploads xlsx without this installed

router = APIRouter(
    prefix="/bulk-leads",
    tags=["bulk-lead-upload"],
)

# -------------------- Schemas --------------------
class BulkUploadResponse(BaseModel):
    total_rows: int
    successful_uploads: int
    failed_uploads: int
    duplicates_skipped: int
    errors: List[Dict[str, Any]]
    uploaded_leads: List[int]
    validation_summary: Dict[str, int]

# -------------------- Helpers --------------------
ALLOWED_EXTS = {".csv", ".xlsx", ".xlsm"}

def _ext(filename: str) -> str:
    return os.path.splitext(filename or "")[1].lower()

def validate_tabular_file(file: Optional[UploadFile]) -> str:
    """
    Validate presence + allowed extension. Returns normalized extension.
    """
    if not file or not file.filename:
        raise HTTPException(status_code=400, detail="File is required.")
    ext = _ext(file.filename)
    if ext not in ALLOWED_EXTS:
        raise HTTPException(status_code=400, detail="Only CSV/XLSX/XLSM files are allowed.")
    if ext in (".xlsx", ".xlsm") and openpyxl is None:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed on the server; cannot parse Excel files."
        )
    return ext

def parse_csv_stream(file: UploadFile) -> List[List[str]]:
    """
    Stream/parse CSV -> list of rows (list[str]).
    """
    file.file.seek(0)
    text_stream = io.TextIOWrapper(file.file, encoding="utf-8", errors="ignore", newline="")
    reader = csv.reader(text_stream)
    return list(reader)

def _to_str(cell_val) -> str:
    """
    Normalize Excel cell values to string (trimmed). None => ''.
    """
    if cell_val is None:
        return ""
    if isinstance(cell_val, str):
        return cell_val.strip()
    # numbers/dates/bools -> string
    return str(cell_val).strip()

def parse_excel_stream(file: UploadFile, sheet_name: Optional[str] = None) -> List[List[str]]:
    """
    Parse Excel (xlsx/xlsm) using openpyxl in read-only mode, return list of list[str].
    Uses the first worksheet unless sheet_name provided.
    """
    file.file.seek(0)
    wb = openpyxl.load_workbook(file.file, read_only=True, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found in workbook.")
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]

    rows: List[List[str]] = []
    for r in ws.iter_rows(values_only=True):
        rows.append([_to_str(v) for v in r])
    return rows

def parse_tabular(file: UploadFile, ext: str, sheet_name: Optional[str]) -> List[List[str]]:
    """
    Unified parser: CSV or Excel -> rows (list[list[str]]).
    """
    if ext == ".csv":
        return parse_csv_stream(file)
    else:
        return parse_excel_stream(file, sheet_name=sheet_name)

def get_column_value(row: List[str], idx: Optional[int]) -> Optional[str]:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    val = (row[idx] or "").strip()
    return val or None

def process_segment(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    parts = [p.strip() for p in text.split(",")]
    parts = [p for p in parts if p]
    return json.dumps(parts) if parts else None

def _bulk_fetch_existing(
    db: Session,
    emails: Set[str],
    mobiles: Set[str],
    pans: Set[str],
) -> Tuple[Set[str], Set[str], Set[str]]:
    """One-shot DB lookups for duplicates."""
    existing_emails: Set[str] = set()
    existing_mobiles: Set[str] = set()
    existing_pans: Set[str] = set()

    if emails:
        for (e,) in db.execute(select(Lead.email).where(Lead.email.in_(emails))).all():
            if e:
                existing_emails.add(e)
    if mobiles:
        for (m,) in db.execute(select(Lead.mobile).where(Lead.mobile.in_(mobiles))).all():
            if m:
                existing_mobiles.add(m)
    if pans:
        for (p,) in db.execute(select(Lead.pan).where(Lead.pan.in_(pans))).all():
            if p:
                existing_pans.add(p.upper())

    return existing_emails, existing_mobiles, existing_pans

# -------------------- Endpoint (FAST + CSV/XLSX) --------------------
@router.post("/upload", response_model=BulkUploadResponse)
async def upload_bulk_leads_fast(
    mobile_column: int = Form(...),

    # optional columns (0-based indices)
    lead_source_id: Optional[int] = Form(None),
    branch_id: Optional[int] = Form(None),
    name_column: Optional[int] = Form(None),
    email_column: Optional[int] = Form(None),
    city_column: Optional[int] = Form(None),
    address_column: Optional[int] = Form(None),
    segment_column: Optional[int] = Form(None),
    occupation_column: Optional[int] = Form(None),
    investment_column: Optional[int] = Form(None),
    pan_column: Optional[int] = Form(None),

    # NEW: support CSV and Excel
    upload_file: UploadFile = File(None),
    sheet_name: Optional[str] = Form(None, description="Excel sheet name (optional)"),

    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    FAST upload: single-pass parse, bulk duplicate checks, and a single bulk insert with RETURNING ids.
    Supports CSV (.csv) and Excel (.xlsx/.xlsm). Column indexes are 0-based.
    """
    employee_code = current_user.employee_code

    # 0) file + foreign validations
    ext = validate_tabular_file(upload_file)

    if lead_source_id is not None:
        exists = db.execute(
            select(func.count()).select_from(LeadSource).where(LeadSource.id == lead_source_id)
        ).scalar() or 0
        if not exists:
            raise HTTPException(status_code=400, detail=f"Lead source with ID {lead_source_id} not found")

    emp_exists = db.execute(
        select(func.count()).select_from(UserDetails).where(UserDetails.employee_code == employee_code)
    ).scalar() or 0
    if not emp_exists:
        raise HTTPException(status_code=400, detail=f"Employee with code {employee_code} not found")

    # 1) parse file -> rows
    try:
        rows = parse_tabular(upload_file, ext, sheet_name)
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=400, detail="Failed to parse file")

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    header, data_rows = rows[0], rows[1:]
    total_rows = len(data_rows)

    # 2) Pre-collect candidate keys to do one-shot duplicate DB queries
    file_emails: Set[str] = set()
    file_mobiles: Set[str] = set()
    file_pans: Set[str] = set()

    for r in data_rows:
        e = get_column_value(r, email_column)
        m = get_column_value(r, mobile_column)
        p = get_column_value(r, pan_column)
        if e:
            file_emails.add(e)
        if m:
            file_mobiles.add(m)
        if p:
            file_pans.add(p.upper())

    existing_emails, existing_mobiles, existing_pans = _bulk_fetch_existing(
        db, file_emails, file_mobiles, file_pans
    )

    # 3) validators & counters
    fmt = FormatValidator()   # your utils (regex compiled once internally)
    errors_list: List[Dict[str, Any]] = []
    uploaded_ids: List[int] = []
    successful_uploads = 0
    failed_uploads = 0
    duplicates_skipped = 0
    validation_summary = {
        "format_errors": 0,
        "duplicate_errors": 0,
        "missing_mobile_errors": 0,
        "database_errors": 0,
    }

    # Track duplicates *within* file itself
    seen_in_file_email: Set[str] = set()
    seen_in_file_mobile: Set[str] = set()
    seen_in_file_pan: Set[str] = set()

    # 4) Build rows for bulk insert (dicts)
    rows_to_insert: List[Dict[str, Any]] = []

    for idx, row in enumerate(data_rows, start=2):  # header is row 1
        try:
            mobile = get_column_value(row, mobile_column)
            name = get_column_value(row, name_column)
            email = get_column_value(row, email_column)
            city = get_column_value(row, city_column)
            address = get_column_value(row, address_column)
            segment_text = get_column_value(row, segment_column)
            occupation = get_column_value(row, occupation_column)
            investment = get_column_value(row, investment_column)
            pan_raw = get_column_value(row, pan_column)
            pan = pan_raw.upper() if pan_raw else None

            # Required per-row
            if not mobile:
                failed_uploads += 1
                validation_summary["missing_mobile_errors"] += 1
                errors_list.append({"row": idx, "errors": ["Missing mobile number"], "data": {"mobile": None}})
                continue

            # 4a) Format checks (fast, in-process)
            fmt_errs = []
            if email and not fmt.validate_email_format(email):
                fmt_errs.append("Invalid email format")
            if mobile and not fmt.validate_mobile_format(mobile):
                fmt_errs.append("Mobile number must be exactly 10 digits")
            if pan and not fmt.validate_pan_format(pan):
                fmt_errs.append("Invalid PAN format. PAN must be in format ABCDE1234F")

            if fmt_errs:
                failed_uploads += 1
                validation_summary["format_errors"] += 1
                errors_list.append({"row": idx, "errors": fmt_errs, "data": {"mobile": mobile, "email": email, "pan": pan}})
                continue

            # 4b) Duplicate checks (DB + in-file)
            dup_errs = []
            if email:
                if email in existing_emails or email in seen_in_file_email:
                    dup_errs.append("Email duplicate")
                else:
                    seen_in_file_email.add(email)
            if mobile:
                if mobile in existing_mobiles or mobile in seen_in_file_mobile:
                    dup_errs.append("Mobile duplicate")
                else:
                    seen_in_file_mobile.add(mobile)
            if pan:
                if pan in existing_pans or pan in seen_in_file_pan:
                    dup_errs.append("PAN duplicate")
                else:
                    seen_in_file_pan.add(pan)

            if dup_errs:
                duplicates_skipped += 1
                validation_summary["duplicate_errors"] += 1
                errors_list.append({"row": idx, "errors": dup_errs, "data": {"mobile": mobile, "email": email, "pan": pan}})
                continue

            # 4c) Prepare row payload
            lead_payload = {
                "mobile": mobile,
                "full_name": name,
                "email": email,
                "city": city,
                "address": address,
                "segment": process_segment(segment_text),
                "occupation": occupation,
                "investment": investment,
                "pan": pan,
                "lead_source_id": lead_source_id,
                "created_by": employee_code,
                "created_by_name": current_user.name,
                "branch_id": branch_id,
            }
            # drop None to keep insert slim
            lead_payload = {k: v for k, v in lead_payload.items() if v is not None}
            rows_to_insert.append(lead_payload)

        except Exception as e:
            failed_uploads += 1
            validation_summary["database_errors"] += 1
            errors_list.append({"row": idx, "errors": [f"Row build error: {str(e)}"]})

    if not rows_to_insert:
        # nothing valid to insert, just return summary
        return BulkUploadResponse(
            total_rows=total_rows,
            successful_uploads=0,
            failed_uploads=failed_uploads,
            duplicates_skipped=duplicates_skipped,
            errors=errors_list,
            uploaded_leads=[],
            validation_summary=validation_summary,
        )

    # 5) BULK INSERT with RETURNING ids (Postgres)
    try:
        # Optional: chunk very large inserts
        BATCH = 5000
        for i in range(0, len(rows_to_insert), BATCH):
            chunk = rows_to_insert[i:i+BATCH]
            stmt = insert(Lead).returning(Lead.id)
            result = db.execute(stmt, chunk)
            new_ids = [row[0] for row in result.fetchall()]
            uploaded_ids.extend(new_ids)
        db.commit()
        successful_uploads = len(uploaded_ids)
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Failed to commit bulk upload: {str(e)}")

    return BulkUploadResponse(
        total_rows=total_rows,
        successful_uploads=successful_uploads,
        failed_uploads=failed_uploads,
        duplicates_skipped=duplicates_skipped,
        errors=errors_list,
        uploaded_leads=uploaded_ids,
        validation_summary=validation_summary,
    )

